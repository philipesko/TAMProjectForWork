from openpyxl import load_workbook
import warnings
import seaborn as sns
import matplotlib.pyplot as plt
from pylab import rcParams
import pandas as pd


warnings.simplefilter('ignore')

wb = load_workbook(filename='TestBook.xlsx', read_only=True)
ws = wb['Sheet1']

for row in ws.rows:
    for cell in row:
        if cell.value is not None:
            print("A")
            #print(cell.value)

totalResult = ws.calculate_dimension()
print(totalResult)
print(wb.sheetnames)
wb.close()

#rcParams['figure.figsize'] = 8, 5

df = pd.read_excel('C:/Users/Philipesko/PycharmProjects/firstProject/TestBook.xlsx')
df.info()
print(df.columns)
print('---------------')
df.describe(include=['object', 'bool'])
test = df.describe(include=['object', 'bool'])





